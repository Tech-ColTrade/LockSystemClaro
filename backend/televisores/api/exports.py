"""Exportaciones a Excel (.xlsx) — sincronizaciones y códigos pin.

Mismo estilo que whaletv: encabezado con el color de marca (F6186A).
"""
from __future__ import annotations

import io

from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from televisores.models import BulkSyncItem, BulkSyncJob, SyncJob, Televisor
from televisores.portal.client import PortalClient, PortalError

XLSX_CONTENT_TYPE = (
    'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
)


def _estilar_encabezado(ws):
    fill = PatternFill('solid', fgColor='F6186A')
    font = Font(bold=True, color='FFFFFF')
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal='center')


def _respuesta_xlsx(wb: Workbook, filename: str) -> HttpResponse:
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    resp = HttpResponse(buffer.getvalue(), content_type=XLSX_CONTENT_TYPE)
    resp['Content-Disposition'] = f'attachment; filename="{filename}"'
    return resp


def _resultado_syncjob(estado: str) -> str:
    return {'terminado': 'Aplicado', 'error': 'Error'}.get(estado, 'En proceso')


def _resultado_item(estado: str) -> str:
    return {'ok': 'Aplicado', 'error': 'Error'}.get(estado, 'Pendiente')


def exportar_sincronizaciones() -> HttpResponse:
    """Historial de cambios de estado (individuales + masivos)."""
    wb = Workbook()
    ws = wb.active
    ws.title = 'Sincronizaciones'
    ws.append(['Fecha', 'Dirección MAC', 'Acción', 'Resultado', 'Tipo', 'Mensaje'])
    _estilar_encabezado(ws)

    filas = []
    for j in SyncJob.objects.select_related('televisor'):
        filas.append((
            j.creado,
            j.televisor.mac_address if j.televisor else '—',
            'Inhabilitar' if j.inhabilitar else 'Habilitar',
            _resultado_syncjob(j.estado),
            'Individual',
            j.error or '',
        ))
    for it in BulkSyncItem.objects.select_related('job'):
        filas.append((
            it.job.creado,
            it.mac_address,
            'Inhabilitar' if it.inhabilitar else 'Habilitar',
            _resultado_item(it.estado),
            'Masivo',
            it.mensaje or '',
        ))

    filas.sort(key=lambda f: f[0], reverse=True)
    for f in filas:
        ws.append([f[0].strftime('%d/%m/%Y %H:%M'), f[1], f[2], f[3], f[4], f[5]])

    for col, ancho in zip('ABCDEF', (18, 20, 14, 12, 12, 40)):
        ws.column_dimensions[col].width = ancho

    return _respuesta_xlsx(wb, 'sincronizaciones.xlsx')


def _txt_bool(valor: bool | None, si: str, no: str) -> str:
    if valor is None:
        return '—'
    return si if valor else no


def exportar_bulk_job(job: BulkSyncJob) -> HttpResponse:
    """Exporta el detalle de UN lote (sincronización o validación masiva)."""
    es_validacion = job.modo == BulkSyncJob.VALIDACION
    wb = Workbook()
    ws = wb.active
    ws.title = 'Validación' if es_validacion else 'Sincronización'

    if es_validacion:
        ws.append(['Dirección MAC', 'Portal', 'App', 'Coincide', 'Mensaje'])
        _estilar_encabezado(ws)
        for it in job.items.all():
            ws.append([
                it.mac_address,
                _txt_bool(it.remoto_inhabilitado, 'Inhabilitado', 'Habilitado'),
                _txt_bool(it.local_inhabilitado, 'Inhabilitado', 'Habilitado'),
                _txt_bool(it.coincide, 'Sí', 'No'),
                it.mensaje or '',
            ])
        anchos = zip('ABCDE', (20, 16, 16, 12, 40))
    else:
        ws.append(['Dirección MAC', 'Acción', 'Resultado', 'Mensaje'])
        _estilar_encabezado(ws)
        for it in job.items.all():
            ws.append([
                it.mac_address,
                'Inhabilitar' if it.inhabilitar else 'Habilitar',
                _resultado_item(it.estado),
                it.mensaje or '',
            ])
        anchos = zip('ABCD', (20, 14, 12, 40))

    for col, ancho in anchos:
        ws.column_dimensions[col].width = ancho

    prefijo = 'validacion_masiva' if es_validacion else 'sincronizacion_masiva'
    return _respuesta_xlsx(wb, f'{prefijo}_{job.pk}.xlsx')


def plantilla_televisores() -> HttpResponse:
    """Plantilla Excel para enrolar televisores."""
    wb = Workbook()
    ws = wb.active
    ws.title = 'Televisores'
    ws.append(['mac_address', 'serial_number', 'numero_credito'])
    _estilar_encabezado(ws)
    ws.append(['B4:04:29:7E:3A:AA', 'B4:04:29:7E:3A:AA', '1234567890'])
    for col, ancho in zip('ABC', (22, 22, 18)):
        ws.column_dimensions[col].width = ancho
    return _respuesta_xlsx(wb, 'plantilla_televisores.xlsx')


def plantilla_estados() -> HttpResponse:
    """Plantilla Excel para enrolar estados (habilitado/inhabilitado)."""
    wb = Workbook()
    ws = wb.active
    ws.title = 'Estados'
    ws.append(['mac_address', 'estado', 'serial_number'])
    _estilar_encabezado(ws)
    ws.append(['B4:04:29:7E:3A:AA', 'inhabilitado', 'B4:04:29:7E:3A:AA'])
    ws.append(['B4:04:29:7E:3A:BB', 'habilitado', 'B4:04:29:7E:3A:BB'])
    for col, ancho in zip('ABC', (22, 16, 22)):
        ws.column_dimensions[col].width = ancho
    return _respuesta_xlsx(wb, 'plantilla_estados.xlsx')


def exportar_pincodes() -> HttpResponse:
    """Códigos Pin disponibles de todos los televisores (consultados al portal)."""
    wb = Workbook()
    ws = wb.active
    ws.title = 'Pin Codes'
    ws.append(['Dirección MAC', 'Código de Acceso', 'Código Pin'])
    _estilar_encabezado(ws)

    client = PortalClient()
    for tv in Televisor.objects.all():
        try:
            grupos = client.get_pin_codes(tv.eui64_portal)
        except (PortalError, ValueError):
            # TV sin registro en el portal o MAC inválida: se omite.
            continue
        for g in grupos:
            ws.append([tv.mac_address, g['passCode'], g['pinCode']])

    for col, ancho in zip('ABC', (20, 18, 16)):
        ws.column_dimensions[col].width = ancho

    return _respuesta_xlsx(wb, 'pincodes.xlsx')

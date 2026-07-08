"""Parseo e importación de televisores desde CSV/XLSX.

Columnas reconocidas (encabezado, insensible a mayúsculas):
    - mac_address   (obligatoria, identifica el televisor)
    - serial_number (opcional)
    - numero_credito(opcional, solo dígitos)
"""
from __future__ import annotations

import csv
import io

from televisores.models import Televisor

COLUMNAS = ('mac_address', 'serial_number', 'numero_credito')


def _normalizar_encabezados(headers: list[str]) -> dict[int, str]:
    """Mapea índice de columna -> nombre de campo reconocido."""
    mapa: dict[int, str] = {}
    for i, h in enumerate(headers):
        clave = (h or '').strip().lower().replace(' ', '_')
        if clave in COLUMNAS:
            mapa[i] = clave
    return mapa


def _filas_csv(data: bytes) -> list[list[str]]:
    texto = data.decode('utf-8-sig', errors='replace')
    return [row for row in csv.reader(io.StringIO(texto))]


def _filas_xlsx(data: bytes) -> list[list[str]]:
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb.active
    filas: list[list[str]] = []
    for row in ws.iter_rows(values_only=True):
        filas.append(['' if c is None else str(c) for c in row])
    wb.close()
    return filas


def leer_filas(nombre: str, data: bytes) -> list[list[str]]:
    """Lee un CSV/XLSX y devuelve sus filas como listas de strings."""
    nombre = (nombre or '').lower()
    if nombre.endswith(('.xlsx', '.xls')):
        return _filas_xlsx(data)
    return _filas_csv(data)


def importar_televisores(nombre: str, data: bytes) -> dict:
    """Crea/actualiza televisores desde el contenido de un archivo.

    Devuelve {'creados', 'actualizados', 'errores': [str, ...]}.
    """
    filas = leer_filas(nombre, data)

    if not filas:
        return {'creados': 0, 'actualizados': 0, 'errores': ['El archivo está vacío.']}

    mapa = _normalizar_encabezados(filas[0])
    if 'mac_address' not in mapa.values():
        return {
            'creados': 0,
            'actualizados': 0,
            'errores': ['Falta la columna obligatoria "mac_address".'],
        }

    creados = 0
    actualizados = 0
    errores: list[str] = []

    for n, fila in enumerate(filas[1:], start=2):
        valores = {campo: (fila[i].strip() if i < len(fila) else '') for i, campo in mapa.items()}
        mac = valores.get('mac_address', '').strip().upper()
        if not mac:
            continue  # fila vacía
        try:
            tv, creado = Televisor.objects.get_or_create(mac_address=mac)
            if 'serial_number' in valores:
                tv.serial_number = valores['serial_number']
            if 'numero_credito' in valores:
                tv.numero_credito = valores['numero_credito']
            tv.full_clean(exclude=['mac_address'])
            tv.save()
            if creado:
                creados += 1
            else:
                actualizados += 1
        except Exception as exc:  # noqa: BLE001 — reportamos por fila sin abortar todo
            errores.append(f'Fila {n} ({mac}): {exc}')

    return {'creados': creados, 'actualizados': actualizados, 'errores': errores}

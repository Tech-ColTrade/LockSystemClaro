"""Constructor de reportes: orígenes con lista blanca de campos.

El usuario elige un ORIGEN (televisores / sincronizaciones / códigos pin) y un
subconjunto de sus CAMPOS. El backend valida TODO contra esta lista blanca y
arma el queryset con el ORM: nunca hay SQL libre ni acceso a campos fuera de los
definidos aquí. Del usuario solo se expone el NOMBRE (nunca email/IP/rol).

Cada origen es "plano" (una fila por registro), así que no hay relaciones que el
usuario pueda armar mal: elige un origen y sus columnas, y ya. Es de solo
lectura y lo puede usar cualquier usuario autenticado.
"""
from __future__ import annotations

import csv
import io
from typing import Callable

from django.db.models import (
    BooleanField,
    Case,
    Count,
    F,
    IntegerField,
    OuterRef,
    Q,
    Subquery,
    Value,
    When,
)
from django.db.models.functions import Coalesce
from django.http import HttpResponse
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from rest_framework import serializers, viewsets
from rest_framework.pagination import PageNumberPagination
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView

from televisores.models import (
    BulkSyncItem,
    BulkSyncJob,
    PinCodeUsado,
    ReporteGuardado,
    SyncJob,
    Televisor,
)

from .exports import _estilar_encabezado, _respuesta_xlsx
from .filtros import filtrar_por_fecha
from .registros import (
    _resultado_item,
    _resultado_syncjob,
    nombre_usuario,
    qs_sincronizaciones,
)


def _fecha(dt) -> str:
    return timezone.localtime(dt).strftime('%d/%m/%Y %H:%M') if dt else '—'


def _mes(dt) -> str:
    return timezone.localtime(dt).strftime('%Y-%m') if dt else '—'


def _dia(dt) -> str:
    return timezone.localtime(dt).strftime('%d/%m/%Y') if dt else '—'


def _dias_desde(dt) -> int:
    """Días transcurridos desde `dt` hasta hoy (antigüedad)."""
    if not dt:
        return 0
    return (timezone.localdate() - timezone.localtime(dt).date()).days


def _tv_etiqueta(serial, mac) -> str:
    """Etiqueta legible de un televisor: su serial, o la MAC si no tiene."""
    s = (serial or '').strip()
    if s and s != '—':
        return s
    return mac or '—'


class Campo:
    """Una columna elegible: su clave, etiqueta, tipo y cómo se renderiza.

    `render` recibe la fila cruda (dict de `.values()`) y devuelve el valor ya
    listo para mostrar/exportar. Así el frontend y el Excel comparten formato.
    `orden` es la columna ORM por la que se puede ordenar (None = no ordenable);
    `orden_invertido` invierte la dirección (p. ej. Antigüedad asc = fecha desc).
    """

    def __init__(
        self,
        key: str,
        label: str,
        tipo: str,
        render: Callable[[dict], object],
        orden: str | None = None,
        orden_invertido: bool = False,
    ):
        self.key = key
        self.label = label
        self.tipo = tipo  # 'texto' | 'fecha' | 'booleano' | 'usuario' | 'numero'
        self.render = render
        self.orden = orden
        self.orden_invertido = orden_invertido


# --- Origen: Televisores ----------------------------------------------------
def _conteo_por_tv(qs_relacion):
    """Subquery con el conteo de filas de una relación para cada televisor.

    Se usa Subquery (y no dos Count sobre relaciones distintas) porque los JOIN
    múltiples multiplican filas y los conteos saldrían inflados.
    """
    sq = (
        qs_relacion.filter(televisor=OuterRef('pk'))
        .order_by()
        .values('televisor')
        .annotate(n=Count('pk'))
        .values('n')
    )
    return Coalesce(Subquery(sq, output_field=IntegerField()), 0)


def _qs_televisores(f: dict):
    qs = Televisor.objects.all()
    q = (f.get('q') or '').strip()
    if q:
        qs = qs.filter(
            Q(serial_number__icontains=q)
            | Q(mac_address__icontains=q)
            | Q(numero_credito__icontains=q)
        )
    inhab = f.get('inhabilitado')
    if inhab in ('true', 'false'):
        qs = qs.filter(inhabilitado=(inhab == 'true'))
    qs = filtrar_por_fecha(qs, f.get('desde'), f.get('hasta'), campo='created_at')
    return (
        qs.annotate(
            # Financiado = tiene número de crédito (mismo criterio del dashboard).
            financiado_bool=Case(
                When(numero_credito='', then=Value(False)),
                default=Value(True),
                output_field=BooleanField(),
            ),
            pines_usados=_conteo_por_tv(PinCodeUsado.objects.all()),
            n_sync_ind=_conteo_por_tv(SyncJob.objects.all()),
            n_sync_bulk=_conteo_por_tv(
                BulkSyncItem.objects.filter(job__modo=BulkSyncJob.SYNC)
            ),
        )
        .annotate(sincronizaciones_n=F('n_sync_ind') + F('n_sync_bulk'))
        .order_by('-created_at')
        .values(
            'serial_number', 'mac_address', 'numero_credito', 'financiado_bool',
            'inhabilitado', 'eui64', 'created_at', 'sincronizaciones_n',
            'pines_usados',
        )
    )


CAMPOS_TELEVISORES = [
    Campo('serial_number', 'Número de serie', 'texto',
          lambda r: r['serial_number'] or '—', orden='serial_number'),
    Campo('mac_address', 'Dirección MAC', 'texto',
          lambda r: r['mac_address'] or '—', orden='mac_address'),
    Campo('numero_credito', 'Número de crédito', 'texto',
          lambda r: r['numero_credito'] or '—', orden='numero_credito'),
    Campo('financiado', 'Financiado', 'booleano',
          lambda r: 'Sí' if r['financiado_bool'] else 'No', orden='financiado_bool'),
    Campo('inhabilitado', 'Estado', 'booleano',
          lambda r: 'Inhabilitado' if r['inhabilitado'] else 'Habilitado',
          orden='inhabilitado'),
    Campo('eui64', 'EUI-64', 'texto', lambda r: r['eui64'] or '—', orden='eui64'),
    Campo('created_at', 'Fecha de registro', 'fecha',
          lambda r: _fecha(r['created_at']), orden='created_at'),
    # Antigüedad asc = registrado más recientemente primero -> fecha desc.
    Campo('antiguedad', 'Antigüedad (días)', 'numero',
          lambda r: _dias_desde(r['created_at']),
          orden='created_at', orden_invertido=True),
    Campo('sincronizaciones', 'Sincronizaciones', 'numero',
          lambda r: r['sincronizaciones_n'], orden='sincronizaciones_n'),
    Campo('pines_usados', 'Pines usados', 'numero',
          lambda r: r['pines_usados'], orden='pines_usados'),
]

# Dimensiones agrupables (modo "agrupado": conteo por grupo). Se reutiliza la
# clase Campo: aquí `render` devuelve la etiqueta del grupo de cada fila.
DIMENSIONES_TELEVISORES = [
    Campo('estado', 'Estado', 'booleano',
          lambda r: 'Inhabilitado' if r['inhabilitado'] else 'Habilitado'),
    Campo('financiado', 'Financiado', 'booleano',
          lambda r: 'Financiado' if r['financiado_bool'] else 'No financiado'),
    Campo('mes_registro', 'Mes de registro', 'fecha', lambda r: _mes(r['created_at'])),
    Campo('dia_registro', 'Día de registro', 'fecha', lambda r: _dia(r['created_at'])),
]


# --- Origen: Sincronizaciones (individuales + masivas) ----------------------
def _qs_sincronizaciones(f: dict):
    # Reutiliza el UNION que ya alimenta la vista /sincronizaciones: filas con
    # las claves fecha, mac, serial, usuario_nombre, inhabilitar, estado, tipo.
    return qs_sincronizaciones(desde=f.get('desde'), hasta=f.get('hasta'))


CAMPOS_SINCRONIZACIONES = [
    Campo('fecha', 'Fecha', 'fecha', lambda r: _fecha(r['fecha']), orden='fecha'),
    Campo('serial_number', 'Número de serie', 'texto',
          lambda r: r['serial'] or '—', orden='serial'),
    Campo('mac_address', 'Dirección MAC', 'texto',
          lambda r: r['mac'] or '—', orden='mac'),
    Campo('usuario', 'Usuario', 'usuario',
          lambda r: r['usuario_nombre'] or '—', orden='usuario_nombre'),
    Campo('accion', 'Acción', 'texto',
          lambda r: 'Inhabilitar' if r['inhabilitar'] else 'Habilitar',
          orden='inhabilitar'),
    # `resultado` no es ordenable: el código de estado significa cosas distintas
    # según el origen (individual vs masivo) y el orden SQL no coincidiría con
    # el texto que se muestra.
    Campo('resultado', 'Resultado', 'texto',
          lambda r: _resultado_syncjob(r['estado'])
          if r['tipo'] == 'Individual' else _resultado_item(r['estado'])),
    Campo('tipo', 'Tipo', 'texto', lambda r: r['tipo'], orden='tipo'),
]

DIMENSIONES_SINCRONIZACIONES = [
    Campo('usuario', 'Usuario', 'usuario', lambda r: r['usuario_nombre'] or '—'),
    # Ranking de equipos: cuántas acciones ha recibido cada televisor.
    Campo('televisor', 'Televisor', 'texto',
          lambda r: _tv_etiqueta(r['serial'], r['mac'])),
    Campo('accion', 'Acción', 'texto',
          lambda r: 'Inhabilitar' if r['inhabilitar'] else 'Habilitar'),
    Campo('resultado', 'Resultado', 'texto',
          lambda r: _resultado_syncjob(r['estado'])
          if r['tipo'] == 'Individual' else _resultado_item(r['estado'])),
    Campo('tipo', 'Tipo', 'texto', lambda r: r['tipo']),
    Campo('mes', 'Mes', 'fecha', lambda r: _mes(r['fecha'])),
    Campo('dia', 'Día', 'fecha', lambda r: _dia(r['fecha'])),
]


# --- Origen: Códigos Pin usados ---------------------------------------------
def _qs_pincodes(f: dict):
    qs = PinCodeUsado.objects.all()
    q = (f.get('q') or '').strip()
    if q:
        qs = qs.filter(
            Q(mac_address__icontains=q)
            | Q(passcode__icontains=q)
            | Q(pin_code__icontains=q)
        )
    qs = filtrar_por_fecha(qs, f.get('desde'), f.get('hasta'), campo='creado')
    return (
        qs.order_by('-creado')
        .annotate(
            usuario_nombre=nombre_usuario('usuario'),
            serial=Coalesce('televisor__serial_number', Value('—')),
        )
        .values('creado', 'serial', 'mac_address', 'usuario_nombre', 'passcode', 'pin_code')
    )


CAMPOS_PINCODES = [
    Campo('fecha', 'Fecha', 'fecha', lambda r: _fecha(r['creado']), orden='creado'),
    Campo('serial_number', 'Número de serie', 'texto',
          lambda r: r['serial'] or '—', orden='serial'),
    Campo('mac_address', 'Dirección MAC', 'texto',
          lambda r: r['mac_address'] or '—', orden='mac_address'),
    Campo('usuario', 'Usuario', 'usuario',
          lambda r: r['usuario_nombre'] or '—', orden='usuario_nombre'),
    Campo('passcode', 'Código de Acceso', 'texto',
          lambda r: r['passcode'], orden='passcode'),
    Campo('pin_code', 'Código Pin', 'texto', lambda r: r['pin_code'], orden='pin_code'),
]

DIMENSIONES_PINCODES = [
    Campo('usuario', 'Usuario', 'usuario', lambda r: r['usuario_nombre'] or '—'),
    # Equipos que más pines consumen.
    Campo('televisor', 'Televisor', 'texto',
          lambda r: _tv_etiqueta(r['serial'], r['mac_address'])),
    Campo('mes', 'Mes', 'fecha', lambda r: _mes(r['creado'])),
    Campo('dia', 'Día', 'fecha', lambda r: _dia(r['creado'])),
]


# --- Registro de orígenes ----------------------------------------------------
ORIGENES = {
    'televisores': {
        'label': 'Televisores',
        'descripcion': 'Una fila por televisor: datos del equipo.',
        'campos': CAMPOS_TELEVISORES,
        'dimensiones': DIMENSIONES_TELEVISORES,
        'filtros': {'fecha': True, 'inhabilitado': True, 'busqueda': True},
        'queryset': _qs_televisores,
        # Desempate al reordenar: sin él, filas "iguales" bailan entre páginas.
        'desempate': ('-pk',),
    },
    'sincronizaciones': {
        'label': 'Sincronizaciones',
        'descripcion': 'Una fila por sincronización (individual o masiva), con quién la hizo.',
        'campos': CAMPOS_SINCRONIZACIONES,
        'dimensiones': DIMENSIONES_SINCRONIZACIONES,
        'filtros': {'fecha': True, 'inhabilitado': False, 'busqueda': False},
        'queryset': _qs_sincronizaciones,
        # Es un UNION: solo puede ordenar/desempatar por columnas del values().
        'desempate': ('tipo', '-id'),
    },
    'pincodes': {
        'label': 'Códigos Pin',
        'descripcion': 'Una fila por cada código pin entregado, con quién lo entregó.',
        'campos': CAMPOS_PINCODES,
        'dimensiones': DIMENSIONES_PINCODES,
        'filtros': {'fecha': True, 'inhabilitado': False, 'busqueda': True},
        'queryset': _qs_pincodes,
        'desempate': ('-pk',),
    },
}


class ReporteInvalido(Exception):
    """La definición del reporte pide un origen o campo fuera de la lista blanca."""


def metadata() -> dict:
    """Descripción de orígenes/campos/filtros para que el frontend pinte la UI."""
    return {
        'origenes': [
            {
                'key': key,
                'label': o['label'],
                'descripcion': o['descripcion'],
                'filtros': o['filtros'],
                'campos': [
                    {'key': c.key, 'label': c.label, 'tipo': c.tipo,
                     'sortable': bool(c.orden)}
                    for c in o['campos']
                ],
                'dimensiones': [
                    {'key': d.key, 'label': d.label, 'tipo': d.tipo}
                    for d in o['dimensiones']
                ],
            }
            for key, o in ORIGENES.items()
        ]
    }


def _validar(origen_key: str, campos_pedidos: list[str]):
    """Resuelve el origen y los campos, rechazando cualquier cosa fuera de la
    lista blanca. Sin campos pedidos, devuelve todos (en su orden natural)."""
    origen = ORIGENES.get(origen_key)
    if origen is None:
        raise ReporteInvalido(f'Origen no válido: «{origen_key}».')
    por_key = {c.key: c for c in origen['campos']}
    if not campos_pedidos:
        return origen, list(origen['campos'])
    seleccion = []
    for k in campos_pedidos:
        campo = por_key.get(k)
        if campo is None:
            raise ReporteInvalido(
                f'Campo no válido para «{origen["label"]}»: «{k}».'
            )
        seleccion.append(campo)
    return origen, seleccion


def _leer_filtros(params) -> dict:
    return {
        'desde': params.get('desde'),
        'hasta': params.get('hasta'),
        'inhabilitado': params.get('inhabilitado'),
        'q': params.get('q'),
    }


def _campos_pedidos(params) -> list[str]:
    raw = params.get('campos') or ''
    return [c for c in (s.strip() for s in raw.split(',')) if c]


def _validar_dimension(origen, dim_key: str):
    """Resuelve la dimensión de agrupación contra la lista blanca del origen."""
    por_key = {d.key: d for d in origen['dimensiones']}
    dim = por_key.get(dim_key)
    if dim is None:
        raise ReporteInvalido(
            f'Agrupación no válida para «{origen["label"]}»: «{dim_key}».'
        )
    return dim


def _agrupar(origen, dimension, filtros) -> list[tuple[str, int]]:
    """Cuenta registros por grupo (GROUP BY + COUNT), en Python.

    Se agrupa en Python y no en SQL porque un origen puede ser un UNION (las
    sincronizaciones), sobre el que anotar un Count es engorroso. El volumen es
    de escala administrativa, así que contar sobre las mismas filas del modo
    lista es simple y correcto. Orden por defecto: cantidad desc, luego etiqueta.
    """
    from collections import Counter

    conteo: Counter = Counter()
    for row in origen['queryset'](filtros):
        conteo[dimension.render(row)] += 1
    return sorted(conteo.items(), key=lambda kv: (-kv[1], str(kv[0])))


ORDENES_AGRUPADO = {'grupo', '-grupo', 'cantidad', '-cantidad'}


def _ordenar_agrupado(items: list[tuple[str, int]], orden: str | None):
    """Reordena los grupos según lo pedido ('' = cantidad desc, el default)."""
    if not orden:
        return items
    if orden not in ORDENES_AGRUPADO:
        raise ReporteInvalido(f'No se puede ordenar por «{orden.lstrip("-")}».')
    desc = orden.startswith('-')
    if orden.endswith('grupo'):
        return sorted(items, key=lambda kv: str(kv[0]).lower(), reverse=desc)
    return sorted(items, key=lambda kv: kv[1], reverse=desc)


def _aplicar_orden(origen, qs, orden: str | None):
    """Reordena el queryset del modo lista por un campo de la lista blanca.

    Prefijo '-' = descendente. Siempre se añade el desempate del origen para
    que la paginación sea estable (filas iguales no bailan entre páginas).
    """
    if not orden:
        return qs
    desc = orden.startswith('-')
    key = orden[1:] if desc else orden
    campo = next((c for c in origen['campos'] if c.key == key), None)
    if campo is None or not campo.orden:
        raise ReporteInvalido(f'No se puede ordenar por «{key}».')
    # XOR: un campo invertido (antigüedad) asciende cuando su columna desciende.
    pfx = '-' if desc != campo.orden_invertido else ''
    return qs.order_by(f'{pfx}{campo.orden}', *origen['desempate'])


def _porcentaje(n: int, total: int) -> float:
    return round(n * 100 / total, 1) if total else 0.0


def _campos_agrupado(dimension) -> list[dict]:
    return [
        {'key': 'grupo', 'label': dimension.label, 'tipo': dimension.tipo,
         'sortable': True},
        {'key': 'cantidad', 'label': 'Cantidad', 'tipo': 'numero', 'sortable': True},
        {'key': 'porcentaje', 'label': '% del total', 'tipo': 'porcentaje',
         'sortable': False},
    ]


def _es_agrupado(params) -> bool:
    return params.get('modo') == 'agrupado'


def _respuesta_csv(filename: str, filas) -> HttpResponse:
    """CSV con separador ';' y BOM UTF-8: así Excel en es-CO lo abre en
    columnas y con tildes correctas al hacer doble clic. Otras herramientas
    (Power BI, pandas) detectan el separador sin problema."""
    buf = io.StringIO()
    # BOM UTF-8 al inicio: sin él, Excel asume Latin-1 y daña las tildes.
    buf.write('﻿')
    w = csv.writer(buf, delimiter=';')
    for fila in filas:
        w.writerow(fila)
    resp = HttpResponse(buf.getvalue(), content_type='text/csv; charset=utf-8')
    resp['Content-Disposition'] = f'attachment; filename="{filename}"'
    return resp


# --- Vistas -----------------------------------------------------------------
class ReportesCamposView(APIView):
    """Metadatos: qué orígenes, campos y filtros existen. Cualquier autenticado."""

    permission_classes = [IsAuthenticated]

    def get(self, request):
        return Response(metadata())


class ReportesConsultarView(APIView):
    """Previsualización paginada: las filas que se van a exportar.

    Modo 'lista' (por defecto): filas planas de los campos elegidos.
    Modo 'agrupado': conteo por la dimensión elegida (+ total general).
    """

    permission_classes = [IsAuthenticated]

    def get(self, request):
        params = request.query_params
        origen = ORIGENES.get(params.get('origen', ''))
        if origen is None:
            return Response(
                {'detail': f'Origen no válido: «{params.get("origen", "")}».'},
                status=400,
            )
        filtros = _leer_filtros(params)
        orden = params.get('orden') or ''
        paginator = PageNumberPagination()

        if _es_agrupado(params):
            try:
                dimension = _validar_dimension(origen, params.get('dimension', ''))
                items = _ordenar_agrupado(
                    _agrupar(origen, dimension, filtros), orden
                )
            except ReporteInvalido as e:
                return Response({'detail': str(e)}, status=400)
            total = sum(n for _, n in items)
            page = paginator.paginate_queryset(items, request, view=self)
            rows = [
                [etiqueta, n, _porcentaje(n, total)] for etiqueta, n in page
            ]
            return Response({
                'count': paginator.page.paginator.count,
                'next': paginator.get_next_link(),
                'previous': paginator.get_previous_link(),
                'campos': _campos_agrupado(dimension),
                'rows': rows,
                'total': total,
            })

        try:
            _, seleccion = _validar(params.get('origen', ''), _campos_pedidos(params))
            qs = _aplicar_orden(origen, origen['queryset'](filtros), orden)
        except ReporteInvalido as e:
            return Response({'detail': str(e)}, status=400)
        page = paginator.paginate_queryset(qs, request, view=self)
        rows = [[c.render(r) for c in seleccion] for r in page]
        return Response({
            'count': paginator.page.paginator.count,
            'next': paginator.get_next_link(),
            'previous': paginator.get_previous_link(),
            'campos': [
                {'key': c.key, 'label': c.label, 'tipo': c.tipo,
                 'sortable': bool(c.orden)}
                for c in seleccion
            ],
            'rows': rows,
            'total': None,
        })


class ReportesExportarView(APIView):
    """Descarga TODAS las filas del reporte.

    `?formato=csv` -> CSV (';' + BOM). Cualquier otro valor -> Excel (.xlsx)
    con el look de marca de siempre.
    """

    permission_classes = [IsAuthenticated]

    def get(self, request):
        params = request.query_params
        origen_key = params.get('origen', '')
        origen = ORIGENES.get(origen_key)
        if origen is None:
            return Response({'detail': f'Origen no válido: «{origen_key}».'}, status=400)
        filtros = _leer_filtros(params)
        orden = params.get('orden') or ''
        es_csv = params.get('formato') == 'csv'

        if _es_agrupado(params):
            try:
                dimension = _validar_dimension(origen, params.get('dimension', ''))
                items = _ordenar_agrupado(
                    _agrupar(origen, dimension, filtros), orden
                )
            except ReporteInvalido as e:
                return Response({'detail': str(e)}, status=400)
            total = sum(n for _, n in items)

            if es_csv:
                filas = [[dimension.label, 'Cantidad', '% del total']]
                filas += [[e, n, _porcentaje(n, total)] for e, n in items]
                filas.append(['Total', total, 100.0 if total else 0])
                return _respuesta_csv(f'reporte_{origen_key}_agrupado.csv', filas)

            wb = Workbook()
            ws = wb.active
            ws.title = 'Reporte'
            ws.append([dimension.label, 'Cantidad', '% del total'])
            _estilar_encabezado(ws)
            for etiqueta, n in items:
                # El % va como número real con formato de porcentaje: en Excel
                # se puede operar con él (no es un texto "64,3 %").
                ws.append([etiqueta, n, (n / total) if total else 0])
            ws.append(['Total', total, 1 if total else 0])
            for fila in ws.iter_rows(min_row=2, min_col=3, max_col=3):
                fila[0].number_format = '0.0%'
            for col, ancho in (('A', 28), ('B', 14), ('C', 12)):
                ws.column_dimensions[col].width = ancho
            return _respuesta_xlsx(wb, f'reporte_{origen_key}_agrupado.xlsx')

        try:
            _, seleccion = _validar(origen_key, _campos_pedidos(params))
            qs = _aplicar_orden(origen, origen['queryset'](filtros), orden)
        except ReporteInvalido as e:
            return Response({'detail': str(e)}, status=400)

        if es_csv:
            filas = [[c.label for c in seleccion]]
            filas += [[c.render(row) for c in seleccion] for row in qs]
            return _respuesta_csv(f'reporte_{origen_key}.csv', filas)

        wb = Workbook()
        ws = wb.active
        ws.title = 'Reporte'
        ws.append([c.label for c in seleccion])
        _estilar_encabezado(ws)
        for row in qs:
            ws.append([c.render(row) for c in seleccion])
        for i in range(1, len(seleccion) + 1):
            ws.column_dimensions[get_column_letter(i)].width = 22
        return _respuesta_xlsx(wb, f'reporte_{origen_key}.xlsx')


# --- Reportes guardados (por usuario) ---------------------------------------
def _validar_definicion(definicion) -> None:
    """Valida una definición guardada contra la lista blanca (mismo criterio
    que las consultas): así no se persiste un origen/campo/dimensión inventado."""
    if not isinstance(definicion, dict):
        raise ReporteInvalido('La definición del reporte es inválida.')
    origen = ORIGENES.get(definicion.get('origen', ''))
    if origen is None:
        raise ReporteInvalido(f'Origen no válido: «{definicion.get("origen", "")}».')
    if definicion.get('modo') == 'agrupado':
        _validar_dimension(origen, definicion.get('dimension', ''))
    else:
        _validar(definicion.get('origen', ''), definicion.get('campos') or [])


def _es_admin(user) -> bool:
    return getattr(user, 'role', '') == 'admin' or bool(user.is_superuser)


class ReporteGuardadoSerializer(serializers.ModelSerializer):
    # `es_propio` deja que el frontend distinga "míos" de "plantillas de otros"
    # (las ajenas se pueden usar pero no editar ni borrar).
    es_propio = serializers.SerializerMethodField()
    creado_por = serializers.SerializerMethodField()

    class Meta:
        model = ReporteGuardado
        fields = ['id', 'nombre', 'definicion', 'compartido', 'es_propio',
                  'creado_por', 'creado']
        read_only_fields = ['id', 'creado']

    def get_es_propio(self, obj) -> bool:
        return obj.usuario_id == self.context['request'].user.pk

    def get_creado_por(self, obj) -> str:
        u = obj.usuario
        if u is None:
            return '—'
        nombre = f'{u.first_name} {u.last_name}'.strip()
        return nombre or u.email

    def validate_nombre(self, value):
        value = (value or '').strip()
        if not value:
            raise serializers.ValidationError('Ponle un nombre.')
        return value

    def validate_definicion(self, value):
        try:
            _validar_definicion(value)
        except ReporteInvalido as e:
            raise serializers.ValidationError(str(e))
        return value

    def validate_compartido(self, value):
        # Compartir como plantilla para todos es potestad del Administrador.
        if value and not _es_admin(self.context['request'].user):
            raise serializers.ValidationError(
                'Solo un administrador puede compartir plantillas.'
            )
        return value

    def validate(self, attrs):
        # Nombre único por usuario (mensaje claro en vez de un 500 por IntegrityError).
        usuario = self.context['request'].user
        nombre = attrs.get('nombre')
        qs = ReporteGuardado.objects.filter(usuario=usuario, nombre=nombre)
        if self.instance:
            qs = qs.exclude(pk=self.instance.pk)
        if nombre and qs.exists():
            raise serializers.ValidationError(
                {'nombre': 'Ya tienes un reporte guardado con ese nombre.'}
            )
        return attrs


class ReporteGuardadoViewSet(viewsets.ModelViewSet):
    """CRUD de los reportes guardados del usuario autenticado.

    Se listan los propios + las plantillas compartidas de otros; editar y
    borrar solo aplica sobre los propios (el queryset de escritura lo acota).
    """

    serializer_class = ReporteGuardadoSerializer
    permission_classes = [IsAuthenticated]
    pagination_class = None  # lista completa: son pocos por usuario

    def get_queryset(self):
        user = self.request.user
        if self.action in ('update', 'partial_update', 'destroy'):
            # Escribir/borrar: solo los míos (una plantilla ajena da 404).
            return ReporteGuardado.objects.filter(usuario=user)
        return ReporteGuardado.objects.filter(
            Q(usuario=user) | Q(compartido=True)
        ).select_related('usuario')

    def perform_create(self, serializer):
        serializer.save(usuario=self.request.user)
